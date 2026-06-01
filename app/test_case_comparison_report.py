"""Markdown comparison report for historical Jira test cases."""
from __future__ import annotations

import logging
import re
from io import BytesIO
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timezone
from statistics import mean
from typing import TYPE_CHECKING, Any, Iterable

from app.jira_ticket_insights import (
    DEFAULT_EXCLUDED_PROJECT_KEYS,
    _TEST_CASE_PATTERNS,
    _include_ticket,
    _is_request_only_test_case_phrase,
    _normalized_project_keys,
    _normalize_space,
    _project_where_clause,
    _scan_ticket,
    _ticket_sources,
)

if TYPE_CHECKING:
    from app.config import Settings


log = logging.getLogger(__name__)


@dataclass(frozen=True)
class CriterionResult:
    name: str
    weight: int
    met: bool
    detail: str


@dataclass(frozen=True)
class TicketQuality:
    ticket_key: str
    project_key: str
    summary: str
    status: str
    issue_type: str
    url: str
    updated_at: str | None
    requirement_docs: list[str]
    test_case_count: int
    score: int
    grade: str
    criteria: list[CriterionResult]
    strengths: list[str]
    gaps: list[str]
    evidence_sources: list[str]
    evidence_snippets: list[str]


@dataclass(frozen=True)
class PipelineComparison:
    ticket_key: str
    project_key: str
    summary: str
    historical_test_case_count: int
    historical_score: int
    historical_grade: str
    generated_test_case_count: int
    generated_score: int
    generated_grade: str
    grounded_repos: list[str]
    semantic_hits_count: int
    files_touched_count: int
    architecture_context_chars: int
    repomix_context_chars: int
    score_delta: int
    case_delta: int
    generated_strengths: list[str]
    generated_gaps: list[str]
    generated_snippet: str
    error: str | None = None


@dataclass(frozen=True)
class ComparisonData:
    rows: list[dict[str, Any]]
    analyzed: list[TicketQuality]
    pipeline_comparisons: list[PipelineComparison]
    project_key: str | None
    excluded_projects: set[str]
    total_with_requirements: int
    limit: int
    pipeline_limit: int


def build_test_case_comparison_report(
    settings: "Settings",
    *,
    project_key: str | None = None,
    limit: int = 500,
    pipeline_limit: int | None = None,
    excluded_project_keys: Iterable[str] | None = None,
) -> tuple[str, str]:
    """Return a filename and Opus-generated Markdown comparison report."""
    data = _collect_comparison_data(
        settings,
        project_key=project_key,
        limit=limit,
        pipeline_limit=pipeline_limit,
        excluded_project_keys=excluded_project_keys,
    )
    deterministic_report = _render_report(
        rows=data.rows,
        analyzed=data.analyzed,
        pipeline_comparisons=data.pipeline_comparisons,
        project_key=data.project_key,
        excluded_projects=data.excluded_projects,
        total_with_requirements=data.total_with_requirements,
        limit=data.limit,
        pipeline_limit=data.pipeline_limit,
    )
    opus_analysis = _generate_opus_analysis(
        settings=settings,
        rows=data.rows,
        analyzed=data.analyzed,
        pipeline_comparisons=data.pipeline_comparisons,
        project_key=data.project_key,
        excluded_projects=data.excluded_projects,
        total_with_requirements=data.total_with_requirements,
        limit=data.limit,
        pipeline_limit=data.pipeline_limit,
    )
    markdown = _compose_opus_report(
        opus_analysis=opus_analysis,
        deterministic_report=deterministic_report,
        model=settings.test_case_comparison_model,
    )
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    suffix = f"-{project_key.strip().upper()}" if project_key else ""
    return f"jira-test-case-quantity-quality-clarity{suffix}-{stamp}.md", markdown


def build_test_case_comparison_sheet(
    settings: "Settings",
    *,
    project_key: str | None = None,
    limit: int = 500,
    pipeline_limit: int | None = None,
    excluded_project_keys: Iterable[str] | None = None,
) -> tuple[str, bytes]:
    """Return a filename and XLSX workbook bytes for spreadsheet comparison."""
    data = _collect_comparison_data(
        settings,
        project_key=project_key,
        limit=limit,
        pipeline_limit=pipeline_limit,
        excluded_project_keys=excluded_project_keys,
    )
    opus_analysis = _generate_opus_analysis(
        settings=settings,
        rows=data.rows,
        analyzed=data.analyzed,
        pipeline_comparisons=data.pipeline_comparisons,
        project_key=data.project_key,
        excluded_projects=data.excluded_projects,
        total_with_requirements=data.total_with_requirements,
        limit=data.limit,
        pipeline_limit=data.pipeline_limit,
    )
    workbook_bytes = _build_xlsx_workbook(data, opus_analysis, settings.test_case_comparison_model)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    suffix = f"-{project_key.strip().upper()}" if project_key else ""
    return f"jira-test-case-comparison{suffix}-{stamp}.xlsx", workbook_bytes


def _collect_comparison_data(
    settings: "Settings",
    *,
    project_key: str | None,
    limit: int,
    pipeline_limit: int | None,
    excluded_project_keys: Iterable[str] | None,
) -> ComparisonData:
    excluded = _normalized_project_keys(
        DEFAULT_EXCLUDED_PROJECT_KEYS if excluded_project_keys is None else excluded_project_keys
    )
    rows = _load_ticket_rows(settings, project_key=project_key, excluded_project_keys=excluded)

    analyzed: list[TicketQuality] = []
    total_with_requirements = 0
    for row in rows:
        insight = _scan_ticket(dict(row), settings.jira_base_url)
        if insight["requirement_docs"]:
            total_with_requirements += 1
        if not _include_ticket(insight, "test_cases"):
            continue
        analyzed.append(_evaluate_ticket(row=dict(row), insight=insight))

    effective_pipeline_limit = _bounded_pipeline_limit(
        pipeline_limit if pipeline_limit is not None else settings.test_case_comparison_pipeline_limit
    )
    pipeline_comparisons = _run_pipeline_comparisons(
        settings=settings,
        analyzed=analyzed,
        rows=rows,
        pipeline_limit=effective_pipeline_limit,
    )
    return ComparisonData(
        rows=rows,
        analyzed=analyzed,
        pipeline_comparisons=pipeline_comparisons,
        project_key=project_key,
        excluded_projects=excluded,
        total_with_requirements=total_with_requirements,
        limit=limit,
        pipeline_limit=effective_pipeline_limit,
    )


def _load_ticket_rows(
    settings: "Settings",
    *,
    project_key: str | None,
    excluded_project_keys: Iterable[str],
) -> list[dict[str, Any]]:
    if not settings.database_url:
        return []

    import psycopg
    from psycopg.rows import dict_row

    where, params = _project_where_clause(project_key, excluded_project_keys)
    with psycopg.connect(settings.database_url, row_factory=dict_row) as conn:
        return [
            dict(row)
            for row in conn.execute(
                f"""
                SELECT ticket_key, project_key, summary, description, status,
                       issue_type, labels, updated_at, fetched_at, data
                FROM jira_ticket_cache
                {where}
                ORDER BY updated_at DESC NULLS LAST, fetched_at DESC NULLS LAST
                """,
                params,
            ).fetchall()
        ]


def _bounded_pipeline_limit(value: Any) -> int:
    try:
        limit = int(value)
    except (TypeError, ValueError):
        limit = 0
    return max(0, min(limit, 50))


def _bounded_top_k(value: Any) -> int:
    try:
        top_k = int(value)
    except (TypeError, ValueError):
        top_k = 15
    return max(1, min(top_k, 50))


def _run_pipeline_comparisons(
    *,
    settings: "Settings",
    analyzed: list[TicketQuality],
    rows: list[dict[str, Any]],
    pipeline_limit: int,
) -> list[PipelineComparison]:
    """Generate actual pipeline test cases for a bounded set of filtered tickets."""
    if pipeline_limit <= 0 or not analyzed:
        return []

    from app.test_case_generator import TestCaseGenerator

    generator = TestCaseGenerator(settings=settings)
    rows_by_key = {str(row.get("ticket_key") or ""): dict(row) for row in rows}
    comparisons: list[PipelineComparison] = []

    for historical in analyzed[:pipeline_limit]:
        row = rows_by_key.get(historical.ticket_key)
        if not row:
            continue

        payload = _ticket_generation_payload(row)
        try:
            generated = generator.generate(
                payload,
                style="plain",
                top_k=_bounded_top_k(settings.test_case_comparison_pipeline_top_k),
            )
            comparisons.append(_evaluate_pipeline_output(historical, generated))
        except Exception as exc:
            log.exception("Pipeline testcase generation failed for %s", historical.ticket_key)
            comparisons.append(_pipeline_error_comparison(historical, str(exc)))

    return comparisons


def _ticket_generation_payload(row: dict[str, Any]) -> dict[str, Any]:
    """Build the ticket input for /testcases/generate without leaking existing test cases when possible."""
    return {
        "issueKey": str(row.get("ticket_key") or ""),
        "key": str(row.get("ticket_key") or ""),
        "summary": str(row.get("summary") or ""),
        "issueType": str(row.get("issue_type") or ""),
        "description": _ticket_generation_description(row),
        "acceptanceCriteria": _ticket_generation_acceptance_criteria(row),
        "labels": _string_list(row.get("labels")),
        "components": _ticket_components(row),
    }


def _ticket_generation_description(row: dict[str, Any]) -> str:
    parts: list[str] = []
    seen: set[tuple[str, str]] = set()
    skip_sources = {"Summary", "Issue type", "Labels", "Components", "Fix versions"}

    for source in _ticket_sources(row):
        if source.source in skip_sources:
            continue
        text = _strip_test_case_evidence(source.text)
        if not text:
            continue
        key = (source.source, text)
        if key in seen:
            continue
        seen.add(key)
        parts.append(f"{source.source}: {text}")

    if not parts:
        fallback = _strip_test_case_evidence(str(row.get("description") or ""))
        if fallback:
            parts.append(f"Description: {fallback}")

    return "\n\n".join(parts)[:12000]


def _ticket_generation_acceptance_criteria(row: dict[str, Any]) -> str:
    criteria: list[str] = []
    for source in _ticket_sources(row):
        text = _strip_test_case_evidence(source.text)
        if not text:
            continue
        if _has(r"\b(acceptance criteria|AC[-\s]?\d+|given\b.+\bwhen\b.+\bthen\b|prd|brd|srs|requirement)\b", text):
            criteria.append(f"{source.source}: {text}")
    return "\n\n".join(criteria)[:8000]


def _strip_test_case_evidence(text: str) -> str:
    """Keep requirement text while avoiding reuse of historical test-case blocks."""
    clean = _normalize_space(str(text or ""))
    if not clean:
        return ""
    clean = re.split(r"\btest[\s-]+cases?\b\s*:?", clean, maxsplit=1, flags=re.IGNORECASE)[0]
    clean = re.split(r"\btestcases?\b\s*:?", clean, maxsplit=1, flags=re.IGNORECASE)[0]
    clean = re.sub(r"(?im)^.*\bTC[-_\s]?\d{1,5}\b.*$", "", clean)
    return _normalize_space(clean)


def _ticket_components(row: dict[str, Any]) -> list[str]:
    ticket = _safe_ticket_data(row.get("data"))
    fields = ticket.get("fields") if isinstance(ticket, dict) else {}
    components = fields.get("components") if isinstance(fields, dict) else []
    return _string_list(_component_names_from_raw(components))


def _safe_ticket_data(value: Any) -> dict[str, Any]:
    if isinstance(value, dict):
        return value
    if isinstance(value, str) and value.strip():
        import json

        try:
            parsed = json.loads(value)
        except json.JSONDecodeError:
            return {}
        return parsed if isinstance(parsed, dict) else {}
    return {}


def _component_names_from_raw(values: Any) -> list[str]:
    if not isinstance(values, list):
        return []
    names: list[str] = []
    for value in values:
        if isinstance(value, str):
            names.append(value)
        elif isinstance(value, dict):
            name = value.get("name") or value.get("value")
            if name:
                names.append(str(name))
    return names


def _string_list(value: Any) -> list[str]:
    if not value:
        return []
    if isinstance(value, list):
        return [str(item) for item in value if str(item).strip()]
    if isinstance(value, tuple):
        return [str(item) for item in value if str(item).strip()]
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return []
        if stripped.startswith("["):
            import json

            try:
                parsed = json.loads(stripped)
            except json.JSONDecodeError:
                parsed = None
            if isinstance(parsed, list):
                return [str(item) for item in parsed if str(item).strip()]
        return [part.strip() for part in stripped.split(",") if part.strip()]
    return [str(value)]


def _evaluate_pipeline_output(historical: TicketQuality, generated: dict[str, Any]) -> PipelineComparison:
    output = str(generated.get("test_cases") or "")
    marker_count = _estimate_test_case_count(output)
    case_count = max(1, marker_count) if output.strip() else 0
    criteria = _quality_criteria(_normalize_space(output), marker_count)
    score = sum(c.weight for c in criteria if c.met)
    strengths = [c.name for c in criteria if c.met]
    gaps = [c.name for c in criteria if not c.met]

    return PipelineComparison(
        ticket_key=historical.ticket_key,
        project_key=historical.project_key,
        summary=historical.summary,
        historical_test_case_count=historical.test_case_count,
        historical_score=historical.score,
        historical_grade=historical.grade,
        generated_test_case_count=case_count,
        generated_score=score,
        generated_grade=_grade(score),
        grounded_repos=[str(repo) for repo in generated.get("grounded_repos") or []],
        semantic_hits_count=int(generated.get("semantic_hits_count") or 0),
        files_touched_count=int(generated.get("files_touched_count") or 0),
        architecture_context_chars=int(generated.get("architecture_context_chars") or 0),
        repomix_context_chars=int(generated.get("repomix_context_chars") or 0),
        score_delta=score - historical.score,
        case_delta=case_count - historical.test_case_count,
        generated_strengths=strengths,
        generated_gaps=gaps,
        generated_snippet=_markdown_excerpt(output),
    )


def _pipeline_error_comparison(historical: TicketQuality, error: str) -> PipelineComparison:
    return PipelineComparison(
        ticket_key=historical.ticket_key,
        project_key=historical.project_key,
        summary=historical.summary,
        historical_test_case_count=historical.test_case_count,
        historical_score=historical.score,
        historical_grade=historical.grade,
        generated_test_case_count=0,
        generated_score=0,
        generated_grade="Failed",
        grounded_repos=[],
        semantic_hits_count=0,
        files_touched_count=0,
        architecture_context_chars=0,
        repomix_context_chars=0,
        score_delta=0 - historical.score,
        case_delta=0 - historical.test_case_count,
        generated_strengths=[],
        generated_gaps=[],
        generated_snippet="",
        error=error[:800],
    )


def _evaluate_ticket(row: dict[str, Any], insight: dict[str, Any]) -> TicketQuality:
    source_texts = _test_case_source_texts(row)
    text = "\n\n".join(text for _, text in source_texts).strip()
    if not text:
        text = "\n\n".join(source.text for source in _ticket_sources(row)).strip()
    normalized = _normalize_space(text)
    marker_count = _estimate_test_case_count(normalized)
    case_count = max(1, marker_count)
    criteria = _quality_criteria(normalized, marker_count)
    score = sum(c.weight for c in criteria if c.met)
    strengths = [c.name for c in criteria if c.met][:5]
    gaps = [c.name for c in criteria if not c.met][:6]

    return TicketQuality(
        ticket_key=str(row.get("ticket_key") or ""),
        project_key=str(row.get("project_key") or ""),
        summary=str(row.get("summary") or ""),
        status=str(row.get("status") or ""),
        issue_type=str(row.get("issue_type") or ""),
        url=str(insight.get("url") or ""),
        updated_at=insight.get("updated_at"),
        requirement_docs=list(insight.get("requirement_docs") or []),
        test_case_count=case_count,
        score=score,
        grade=_grade(score),
        criteria=criteria,
        strengths=strengths,
        gaps=gaps,
        evidence_sources=sorted({source for source, _ in source_texts}) or ["Ticket fields"],
        evidence_snippets=_evidence_snippets(source_texts),
    )


def _test_case_source_texts(row: dict[str, Any]) -> list[tuple[str, str]]:
    sources = _ticket_sources(row)
    selected: list[tuple[str, str]] = []
    issue_type_is_test_case = False
    for source in sources:
        text = _normalize_space(source.text)
        if not text:
            continue
        if source.source == "Issue type" and _source_has_test_case_signal(text):
            issue_type_is_test_case = True
        if _source_has_test_case_signal(text):
            selected.append((source.source, text))

    if issue_type_is_test_case:
        for source in sources:
            text = _normalize_space(source.text)
            if text and (source.source, text) not in selected:
                selected.append((source.source, text))
    elif selected:
        for source in sources:
            text = _normalize_space(source.text)
            if text and (source.source, text) not in selected:
                selected.append((source.source, text))
    return selected


def _source_has_test_case_signal(text: str) -> bool:
    for pattern in _TEST_CASE_PATTERNS:
        for match in pattern.finditer(text):
            if not _is_request_only_test_case_phrase(text, match.start()):
                return True
    return False


def _quality_criteria(text: str, case_count: int) -> list[CriterionResult]:
    coverage_terms = {
        "positive": r"\b(positive|happy path|success)\b",
        "negative": r"\b(negative|error|failure|invalid|malformed)\b",
        "edge": r"\b(edge|boundary|limit|empty|null|pagination|date range)\b",
        "regression": r"\b(regression|reproduction|fix verification)\b",
        "security": r"\b(security|auth|authorization|permission|role)\b",
    }
    coverage_hits = [name for name, pattern in coverage_terms.items() if _has(pattern, text)]
    has_expected = _has(r"\b(expected result|expected outcome|then\b|should return|returns? status|result:)", text)
    ambiguous = _has(r"\b(TBD|N/A|as expected|etc\.?|maybe|should work|valid input|either)\b", text)
    if " or " in text.lower() and has_expected:
        ambiguous = True

    return [
        CriterionResult(
            "Identifiable test-case structure",
            10,
            case_count >= 1,
            f"{case_count} case marker(s) found",
        ),
        CriterionResult(
            "Step-by-step execution flow",
            12,
            _has(r"\b(steps?:|given\b|when\b|scenario:|\n\s*\d+\.)", text),
            "Steps, numbered flow, or Gherkin flow is present",
        ),
        CriterionResult(
            "Expected results are stated",
            12,
            has_expected,
            "Expected outcome/result language is present",
        ),
        CriterionResult(
            "Concrete assertions or verification points",
            12,
            _has(r"\b(assertions?:|assert|verify|validate|status code|response body|row count|database row)", text),
            "Assertions, validation points, or observable checks are present",
        ),
        CriterionResult(
            "Coverage mix beyond happy path",
            10,
            len(coverage_hits) >= 2,
            f"Coverage signals: {', '.join(coverage_hits) if coverage_hits else 'none'}",
        ),
        CriterionResult(
            "Traceability to ticket or acceptance criteria",
            10,
            _has(r"\b(ticket reference|acceptance criteria|AC[-\s]?\d+|requirement|JIRA|covers)\b", text),
            "Ticket, AC, or requirement reference is present",
        ),
        CriterionResult(
            "Test data, fixtures, and preconditions",
            10,
            _has(r"\b(test data|payload|fixture|precondition|DB fixtures?|auth fixtures?|mock|setup)\b", text),
            "Data/setup requirements are present",
        ),
        CriterionResult(
            "Source or implementation grounding",
            12,
            _has(r"\b(source references?|route|endpoint|controller|service|module|table|\.py\b|\.js\b|/api/|file:)", text),
            "Code, route, table, or file reference is present",
        ),
        CriterionResult(
            "Automation readiness",
            6,
            _has(r"\b(automation notes?|pytest|JUnit|Postman|Cypress|Playwright|Jest|manual|automated)\b", text),
            "Automation or execution mode is stated",
        ),
        CriterionResult(
            "Unambiguous expected contract",
            6,
            has_expected and not ambiguous,
            "Expected result avoids TBD/N/A/either/or style ambiguity",
        ),
    ]


def _estimate_test_case_count(text: str) -> int:
    patterns = [
        r"^###\s+TC[-_\s]?\d{1,5}",
        r"\bTC[-_\s]?\d{1,5}\b",
        r"\btest[\s-]+case\s*#?\d{1,5}\b",
        r"\bscenario:\s+",
    ]
    counts = [len(re.findall(pattern, text, flags=re.IGNORECASE | re.MULTILINE)) for pattern in patterns]
    return max(counts) if counts else 0


_QUALITY_CRITERIA = {
    "Concrete assertions or verification points",
    "Coverage mix beyond happy path",
    "Traceability to ticket or acceptance criteria",
    "Test data, fixtures, and preconditions",
    "Source or implementation grounding",
    "Automation readiness",
}

_CLARITY_CRITERIA = {
    "Identifiable test-case structure",
    "Step-by-step execution flow",
    "Expected results are stated",
    "Unambiguous expected contract",
}


def _quantity_summary(analyzed: list[TicketQuality]) -> dict[str, Any]:
    if not analyzed:
        return {
            "avg_cases_per_ticket": 0,
            "tickets_at_or_above_pipeline_min": 0,
        }
    case_counts = [item.test_case_count for item in analyzed]
    return {
        "avg_cases_per_ticket": round(mean(case_counts), 1),
        "tickets_at_or_above_pipeline_min": sum(1 for count in case_counts if count >= 6),
    }


def _dimension_pass_rate(analyzed: list[TicketQuality], criterion_names: set[str]) -> float:
    if not analyzed or not criterion_names:
        return 0
    total = 0
    passed = 0
    for ticket in analyzed:
        for criterion in ticket.criteria:
            if criterion.name not in criterion_names:
                continue
            total += 1
            if criterion.met:
                passed += 1
    return round((passed / total) * 100, 1) if total else 0


def _generate_opus_analysis(
    *,
    settings: "Settings",
    rows: list[dict[str, Any]],
    analyzed: list[TicketQuality],
    pipeline_comparisons: list[PipelineComparison],
    project_key: str | None,
    excluded_projects: set[str],
    total_with_requirements: int,
    limit: int,
    pipeline_limit: int,
) -> str:
    if not settings.anthropic_api_key:
        raise RuntimeError("ANTHROPIC_API_KEY is required to generate the comparative analysis with Anthropic Opus")

    try:
        from anthropic import Anthropic
    except ImportError as exc:
        raise RuntimeError("The 'anthropic' package is required to generate the comparative analysis") from exc

    prompt = _opus_user_prompt(
        rows=rows,
        analyzed=analyzed,
        pipeline_comparisons=pipeline_comparisons,
        project_key=project_key,
        excluded_projects=excluded_projects,
        total_with_requirements=total_with_requirements,
        limit=limit,
        pipeline_limit=pipeline_limit,
    )
    client = Anthropic(
        api_key=settings.anthropic_api_key,
        timeout=settings.llm_test_case_timeout_seconds,
    )
    response = client.messages.create(
        model=settings.test_case_comparison_model,
        max_tokens=settings.test_case_comparison_max_tokens,
        system=_OPUS_SYSTEM_PROMPT,
        messages=[{"role": "user", "content": prompt}],
        temperature=0.1,
    )
    text = "".join(
        block.text
        for block in response.content
        if getattr(block, "type", None) == "text" and getattr(block, "text", None)
    ).strip()
    if not text:
        raise RuntimeError("Anthropic Opus returned an empty comparative analysis")
    return text


_OPUS_SYSTEM_PROMPT = """You are a senior QA engineering leader writing an evidence-based comparative analysis.
You compare historical Jira test cases against a local automated test-case generation pipeline.

Requirements:
- Output Markdown only.
- Start with `## Quality, Quantity, and Clarity Comparison`.
- Be candid and executive-readable, but keep claims grounded in the supplied metrics.
- Do not invent ticket counts, projects, or examples beyond the supplied data.
- Compare only these dimensions: quantity, quality, and clarity.
- Explain what already-present Jira ticket test cases provide versus what the new pipeline can create.
- When actual /testcases/generate results are supplied, use them as the primary pipeline evidence.
- Mention that AIGOV is excluded when it appears in the supplied excluded projects.
"""


def _opus_user_prompt(
    *,
    rows: list[dict[str, Any]],
    analyzed: list[TicketQuality],
    pipeline_comparisons: list[PipelineComparison],
    project_key: str | None,
    excluded_projects: set[str],
    total_with_requirements: int,
    limit: int,
    pipeline_limit: int,
) -> str:
    scores = [item.score for item in analyzed]
    grade_counts = Counter(item.grade for item in analyzed)
    avg_score = round(mean(scores), 1) if scores else 0
    scanned_count = len(rows)
    test_case_count = len(analyzed)
    coverage = round((test_case_count / scanned_count) * 100, 1) if scanned_count else 0
    quantity = _quantity_summary(analyzed)
    quality_rate = _dimension_pass_rate(analyzed, _QUALITY_CRITERIA)
    clarity_rate = _dimension_pass_rate(analyzed, _CLARITY_CRITERIA)
    criterion_rows = _criterion_summary(analyzed)
    pipeline_summary = _pipeline_summary(pipeline_comparisons)
    low_items = sorted(analyzed, key=lambda item: item.score)[:25]
    high_items = sorted(analyzed, key=lambda item: item.score, reverse=True)[:10]

    lines = [
        "Create a comparative analysis report from this source data.",
        "",
        "Scope:",
        f"- Jira scope: {project_key.strip().upper() if project_key else 'All cached Jira projects'}",
        f"- Excluded Jira spaces: {', '.join(sorted(excluded_projects)) or 'None'}",
        f"- Detailed ticket rows requested for appendix: {limit}",
        "",
        "Aggregate metrics:",
        f"- Jira tickets scanned: {scanned_count}",
        f"- Tickets containing test-case evidence: {test_case_count}",
        f"- Test-case coverage across scanned tickets: {coverage}%",
        f"- Tickets mentioning PRD / BRD / SRS: {total_with_requirements}",
        f"- Average historical quality score: {avg_score}/100",
        f"- Grade distribution: High={grade_counts.get('High', 0)}, Medium={grade_counts.get('Medium', 0)}, Low={grade_counts.get('Low', 0)}",
        "",
        "Quantity comparison data:",
        f"- Existing ticket test cases: average estimated cases per ticket = {quantity['avg_cases_per_ticket']}",
        f"- Existing tickets with 6+ estimated cases = {quantity['tickets_at_or_above_pipeline_min']}/{test_case_count}",
        "- New pipeline target = 6-8 generated test cases per ticket in plain QA format",
        f"- Actual generated pipeline sample = {pipeline_summary['successful']}/{pipeline_summary['attempted']} tickets succeeded, capped at {pipeline_limit}",
        f"- Actual generated average cases per successful ticket = {pipeline_summary['avg_generated_cases']}",
        f"- Actual generated average case-count delta vs same tickets = {pipeline_summary['avg_case_delta']}",
        "",
        "Quality comparison data:",
        f"- Existing average quality score = {avg_score}/100",
        f"- Existing quality-dimension pass rate = {quality_rate}%",
        "- New pipeline target = source-grounded, traceable cases with happy/edge/failure coverage, fixtures, assertions, and automation notes",
        f"- Actual generated average quality score = {pipeline_summary['avg_generated_score']}/100",
        f"- Actual generated average score delta vs same tickets = {pipeline_summary['avg_score_delta']}",
        "",
        "Clarity comparison data:",
        f"- Existing clarity-dimension pass rate = {clarity_rate}%",
        "- New pipeline target = stable TC identifiers, step-by-step flow, one expected result, and unambiguous contract per case",
        f"- Actual generated clarity-dimension pass rate = {pipeline_summary['generated_clarity_rate']}%",
        "",
        "Actual /testcases/generate comparison rows:",
        "| Ticket | Project | Detected repos | Historical cases | Generated cases | Historical score | Generated score | Score delta | Generated gaps / error |",
        "|---|---|---|---:|---:|---:|---:|---:|---|",
    ]
    for item in pipeline_comparisons:
        lines.append(
            "| "
            + " | ".join(
                [
                    _md_cell(item.ticket_key),
                    _md_cell(item.project_key),
                    _md_cell(", ".join(item.grounded_repos) or "-"),
                    str(item.historical_test_case_count),
                    str(item.generated_test_case_count),
                    str(item.historical_score),
                    str(item.generated_score),
                    str(item.score_delta),
                    _md_cell(item.error or "; ".join(item.generated_gaps) or "-"),
                ]
            )
            + " |"
        )

    lines.extend([
        "",
        "Pipeline benchmark facts:",
        "- The pipeline uses ticket summary, description, issue type, acceptance criteria, labels, and components.",
        "- It auto-detects relevant repositories, with explicit repo override support.",
        "- It retrieves Qdrant semantic code hits and targeted Repomix packed-file context.",
        "- It grounds generated cases in actual routes, files, modules, services, middleware, tables, or snippets.",
        "- It produces 6-8 production-ready plain QA cases, or Gherkin, pytest, or JUnit.",
        "- It requires happy path, edge, and failure/error coverage.",
        "- It calls out auth, authorization, validation, persistence, fixtures, expected results, assertions, automation notes, and source references.",
        "- It uses Gaps/Open Questions rather than inventing behavior when code context is insufficient.",
        "",
        "Historical pass rates by pipeline criterion:",
        "| Criterion | Pass rate | Passing tickets | Comparison gap |",
        "|---|---:|---:|---|",
    ])
    for item in criterion_rows:
        lines.append(f"| {item['name']} | {item['pass_rate']}% | {item['passed']}/{item['total']} | {item['gap']} |")

    lines.extend([
        "",
        "Lowest-scoring historical tickets:",
        "| Ticket | Project | Score | Grade | Strengths | Gaps |",
        "|---|---|---:|---|---|---|",
    ])
    for item in low_items:
        lines.append(
            "| "
            + " | ".join(
                [
                    _md_cell(item.ticket_key),
                    _md_cell(item.project_key),
                    str(item.score),
                    item.grade,
                    _md_cell("; ".join(item.strengths) or "-"),
                    _md_cell("; ".join(item.gaps) or "-"),
                ]
            )
            + " |"
        )

    lines.extend([
        "",
        "Best historical examples:",
        "| Ticket | Project | Score | Grade | Strengths |",
        "|---|---|---:|---|---|",
    ])
    for item in high_items:
        lines.append(
            "| "
            + " | ".join(
                [
                    _md_cell(item.ticket_key),
                    _md_cell(item.project_key),
                    str(item.score),
                    item.grade,
                    _md_cell("; ".join(item.strengths) or "-"),
                ]
            )
            + " |"
        )

    lines.extend([
        "",
        "Write exactly these sections:",
        "1. Executive summary",
        "2. Quantity comparison",
        "3. Quality comparison",
        "4. Clarity comparison",
        "5. Bottom line",
    ])
    return "\n".join(lines)


def _compose_opus_report(*, opus_analysis: str, deterministic_report: str, model: str) -> str:
    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    return "\n".join(
        [
            "# Test Case Quantity, Quality, and Clarity Comparison",
            "",
            f"Generated at: {generated_at}",
            f"Generated narrative model: Anthropic Opus (`{model}`)",
            "",
            opus_analysis.strip(),
            "",
            "---",
            "",
            "## Deterministic Quantity, Quality, and Clarity Appendix",
            "",
            deterministic_report.strip(),
            "",
        ]
    )


def _build_xlsx_workbook(data: ComparisonData, opus_analysis: str, model: str) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise RuntimeError("The 'openpyxl' package is required to generate the Excel comparison sheet") from exc

    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    muted_fill = PatternFill("solid", fgColor="E5E7EB")

    summary = _comparison_summary_values(data)

    ws = wb.active
    ws.title = "Summary"
    _write_rows(
        ws,
        ["Metric", "Value"],
        [[key, value] for key, value in summary.items()],
        header_fill=header_fill,
        header_font=header_font,
    )
    ws["A1"].fill = muted_fill
    ws["B1"].fill = muted_fill

    comparison_rows = [
        [
            "Quantity",
            f"{summary['Tickets with test-case evidence']} tickets, average {summary['Historical avg cases per ticket']} cases",
            f"Actual sample average {summary['Pipeline avg generated cases']} generated cases; target 6-8",
        ],
        [
            "Quality",
            f"Average historical score {summary['Historical avg score']}/100",
            f"Actual sample average generated score {summary['Pipeline avg generated score']}/100",
        ],
        [
            "Clarity",
            f"Historical clarity pass rate {summary['Historical clarity pass rate']}%",
            f"Generated clarity pass rate {summary['Pipeline clarity pass rate']}%",
        ],
    ]
    ws = wb.create_sheet("QQC Comparison")
    _write_rows(
        ws,
        ["Dimension", "Already present in Jira tickets", "Generated by pipeline"],
        comparison_rows,
        header_fill=header_fill,
        header_font=header_font,
    )

    ws = wb.create_sheet("Pipeline Comparison")
    _write_rows(
        ws,
        [
            "Ticket",
            "Project",
            "Summary",
            "Detected repos",
            "Existing cases",
            "Generated cases",
            "Case delta",
            "Existing score",
            "Generated score",
            "Score delta",
            "Existing grade",
            "Generated grade",
            "Semantic hits",
            "Files touched",
            "Architecture chars",
            "Repomix chars",
            "Generated strengths",
            "Generated gaps",
            "Error",
        ],
        [
            [
                item.ticket_key,
                item.project_key,
                item.summary,
                ", ".join(item.grounded_repos),
                item.historical_test_case_count,
                item.generated_test_case_count,
                item.case_delta,
                item.historical_score,
                item.generated_score,
                item.score_delta,
                item.historical_grade,
                item.generated_grade,
                item.semantic_hits_count,
                item.files_touched_count,
                item.architecture_context_chars,
                item.repomix_context_chars,
                "; ".join(item.generated_strengths),
                "; ".join(item.generated_gaps),
                item.error or "",
            ]
            for item in data.pipeline_comparisons
        ],
        header_fill=header_fill,
        header_font=header_font,
    )

    ws = wb.create_sheet("Historical Tickets")
    _write_rows(
        ws,
        [
            "Ticket",
            "Project",
            "Summary",
            "Status",
            "Issue type",
            "Updated",
            "Requirement docs",
            "Estimated cases",
            "Score",
            "Grade",
            "Strengths",
            "Gaps",
            "Evidence sources",
            "Evidence snippets",
            "URL",
        ],
        [
            [
                item.ticket_key,
                item.project_key,
                item.summary,
                item.status,
                item.issue_type,
                item.updated_at or "",
                ", ".join(item.requirement_docs),
                item.test_case_count,
                item.score,
                item.grade,
                "; ".join(item.strengths),
                "; ".join(item.gaps),
                ", ".join(item.evidence_sources),
                "\n".join(item.evidence_snippets),
                item.url,
            ]
            for item in data.analyzed[: data.limit]
        ],
        header_fill=header_fill,
        header_font=header_font,
    )

    ws = wb.create_sheet("Rubric Pass Rates")
    _write_rows(
        ws,
        ["Criterion", "Weight", "Historical pass rate", "Historical passing", "Pipeline expectation"],
        [
            [
                item["name"],
                _criterion_weight(str(item["name"])),
                item["pass_rate"],
                f"{item['passed']}/{item['total']}",
                item["gap"],
            ]
            for item in _criterion_summary(data.analyzed)
        ],
        header_fill=header_fill,
        header_font=header_font,
    )

    ws = wb.create_sheet("Generated Evidence")
    _write_rows(
        ws,
        ["Ticket", "Detected repos", "Generated strengths", "Generated gaps", "Generated snippet", "Error"],
        [
            [
                item.ticket_key,
                ", ".join(item.grounded_repos),
                "; ".join(item.generated_strengths),
                "; ".join(item.generated_gaps),
                item.generated_snippet,
                item.error or "",
            ]
            for item in data.pipeline_comparisons
        ],
        header_fill=header_fill,
        header_font=header_font,
    )

    ws = wb.create_sheet("Opus Narrative")
    _write_rows(
        ws,
        ["Field", "Value"],
        [
            ["Model", model],
            ["Narrative", opus_analysis],
        ],
        header_fill=header_fill,
        header_font=header_font,
    )
    ws["B2"].alignment = Alignment(wrap_text=True, vertical="top")

    for worksheet in wb.worksheets:
        _format_sheet(worksheet)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def _comparison_summary_values(data: ComparisonData) -> dict[str, Any]:
    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    scores = [item.score for item in data.analyzed]
    grade_counts = Counter(item.grade for item in data.analyzed)
    scanned_count = len(data.rows)
    test_case_count = len(data.analyzed)
    coverage = round((test_case_count / scanned_count) * 100, 1) if scanned_count else 0
    quantity = _quantity_summary(data.analyzed)
    pipeline = _pipeline_summary(data.pipeline_comparisons)
    return {
        "Generated at": generated_at,
        "Scope": data.project_key.strip().upper() if data.project_key else "All cached Jira projects",
        "Excluded Jira spaces": ", ".join(sorted(data.excluded_projects)) or "None",
        "Jira tickets scanned": scanned_count,
        "Tickets with test-case evidence": test_case_count,
        "Ticket test-case coverage %": coverage,
        "Tickets mentioning PRD / BRD / SRS": data.total_with_requirements,
        "Historical avg cases per ticket": quantity["avg_cases_per_ticket"],
        "Historical tickets with 6+ cases": quantity["tickets_at_or_above_pipeline_min"],
        "Historical avg score": round(mean(scores), 1) if scores else 0,
        "Historical high grade tickets": grade_counts.get("High", 0),
        "Historical medium grade tickets": grade_counts.get("Medium", 0),
        "Historical low grade tickets": grade_counts.get("Low", 0),
        "Historical quality pass rate": _dimension_pass_rate(data.analyzed, _QUALITY_CRITERIA),
        "Historical clarity pass rate": _dimension_pass_rate(data.analyzed, _CLARITY_CRITERIA),
        "Pipeline sample cap": data.pipeline_limit,
        "Pipeline attempted": pipeline["attempted"],
        "Pipeline successful": pipeline["successful"],
        "Pipeline failed": pipeline["failed"],
        "Pipeline avg generated cases": pipeline["avg_generated_cases"],
        "Pipeline avg case delta": pipeline["avg_case_delta"],
        "Pipeline avg generated score": pipeline["avg_generated_score"],
        "Pipeline avg score delta": pipeline["avg_score_delta"],
        "Pipeline quality pass rate": pipeline["generated_quality_rate"],
        "Pipeline clarity pass rate": pipeline["generated_clarity_rate"],
    }


def _criterion_weight(name: str) -> int:
    for criterion_name, weight, _ in _rubric_rows():
        if criterion_name == name:
            return weight
    return 0


def _write_rows(
    ws: Any,
    headers: list[str],
    rows: list[list[Any]],
    *,
    header_fill: Any,
    header_font: Any,
) -> None:
    from openpyxl.styles import Alignment

    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

    for row in rows:
        ws.append([_xlsx_cell_value(value) for value in row])

    if headers:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions


def _format_sheet(ws: Any) -> None:
    from openpyxl.styles import Alignment

    wrap_keywords = ("summary", "snippet", "evidence", "strength", "gap", "error", "narrative", "expectation")
    for column_cells in ws.columns:
        header = str(column_cells[0].value or "").lower()
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        width = min(max(max_len + 2, 12), 70)
        ws.column_dimensions[column_cells[0].column_letter].width = width
        should_wrap = any(keyword in header for keyword in wrap_keywords) or width >= 50
        for cell in column_cells:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=should_wrap,
            )


def _xlsx_cell_value(value: Any) -> Any:
    if isinstance(value, (int, float)) or value is None:
        return value
    text = _clean_xlsx_text(str(value))
    if text.startswith(("=", "+", "-", "@")):
        text = "'" + text
    return text


def _clean_xlsx_text(value: str) -> str:
    return re.sub(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]", "", value)


def _render_report(
    *,
    rows: list[dict[str, Any]],
    analyzed: list[TicketQuality],
    pipeline_comparisons: list[PipelineComparison],
    project_key: str | None,
    excluded_projects: set[str],
    total_with_requirements: int,
    limit: int,
    pipeline_limit: int,
) -> str:
    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    scores = [item.score for item in analyzed]
    grade_counts = Counter(item.grade for item in analyzed)
    avg_score = round(mean(scores), 1) if scores else 0
    scanned_count = len(rows)
    test_case_count = len(analyzed)
    shown_count = min(test_case_count, limit)
    coverage = round((test_case_count / scanned_count) * 100, 1) if scanned_count else 0
    quantity = _quantity_summary(analyzed)
    quality_rate = _dimension_pass_rate(analyzed, _QUALITY_CRITERIA)
    clarity_rate = _dimension_pass_rate(analyzed, _CLARITY_CRITERIA)
    criterion_rows = _criterion_summary(analyzed)
    pipeline_summary = _pipeline_summary(pipeline_comparisons)

    lines = [
        "# Test Case Quantity, Quality, and Clarity Comparison",
        "",
        f"Generated at: {generated_at}",
        f"Scope: {project_key.strip().upper() if project_key else 'All cached Jira projects'}",
        f"Excluded Jira spaces: {', '.join(sorted(excluded_projects)) or 'None'}",
        "",
        "## Executive Summary",
        "",
        f"- Jira tickets scanned: {scanned_count}",
        f"- Quantity: {test_case_count} tickets contain test-case evidence ({coverage}% of scanned tickets). Existing tickets average {quantity['avg_cases_per_ticket']} estimated case(s) each; the new pipeline creates 6-8 cases per ticket.",
        f"- Quality: existing ticket test cases average {avg_score}/100 with distribution High {grade_counts.get('High', 0)}, Medium {grade_counts.get('Medium', 0)}, Low {grade_counts.get('Low', 0)}. The new pipeline targets source-grounded, traceable, fixture-backed cases with broad coverage.",
        f"- Clarity: existing ticket test cases pass {clarity_rate}% of clarity checks. The new pipeline targets stable TC identifiers, steps, expected results, and unambiguous assertions.",
        f"- Actual pipeline sample: {pipeline_summary['successful']}/{pipeline_summary['attempted']} filtered tickets generated successfully via `/testcases/generate` (cap {pipeline_limit}).",
        f"- PRD / BRD / SRS context found in {total_with_requirements} scanned tickets.",
        f"- Tickets included in detailed section: {shown_count}",
        "",
        "## Quantity, Quality, and Clarity Comparison",
        "",
        "| Dimension | Already present in Jira tickets | New pipeline can create |",
        "|---|---|---|",
        f"| Quantity | {test_case_count} tickets with test-case evidence; average {quantity['avg_cases_per_ticket']} estimated case(s) per ticket; {quantity['tickets_at_or_above_pipeline_min']} tickets have at least 6 estimated cases. | Actual sample average {pipeline_summary['avg_generated_cases']} generated case(s) per successful ticket; target is 6-8 production-ready cases per ticket. |",
        f"| Quality | Average score {avg_score}/100; quality-dimension pass rate {quality_rate}%. | Actual sample average score {pipeline_summary['avg_generated_score']}/100, average delta {pipeline_summary['avg_score_delta']} against the same tickets; source-grounded cases with happy, edge, failure, validation, auth, persistence, fixtures, assertions, automation notes, and ticket traceability. |",
        f"| Clarity | Clarity-dimension pass rate {clarity_rate}%. | Actual sample clarity pass rate {pipeline_summary['generated_clarity_rate']}%; stable TC IDs/headings, step-by-step flow, one expected result, explicit assertions, and gap/open-question handling when behavior is unknown. |",
        "",
        "## Pipeline Baseline",
        "",
        "The local test-case pipeline creates cases by:",
        "",
        "- Takes ticket summary, description, issue type, acceptance criteria, labels, and components as input.",
        "- Detects relevant repositories automatically, with an explicit repo override when needed.",
        "- Retrieves Qdrant semantic code hits and targeted Repomix packed-file context.",
        "- Grounds every generated case in real routes, files, modules, services, middleware, tables, or snippets.",
        "- Produces 6-8 production-ready plain QA cases, or Gherkin, pytest, or JUnit styles.",
        "- Requires happy path, edge cases, and failure/error coverage.",
        "- Calls out authentication, authorization, validation, and persistence checks when relevant.",
        "- Requires concrete payloads, expected results, assertions, fixtures, automation notes, and ticket traceability.",
        "- Uses a Gaps/Open Questions section rather than inventing behavior when code context is insufficient.",
        "",
        "## Scoring Rubric",
        "",
        "| Criterion | Weight | Pipeline expectation |",
        "|---|---:|---|",
    ]
    for name, weight, expectation in _rubric_rows():
        lines.append(f"| {name} | {weight} | {expectation} |")

    lines.extend([
        "",
        "## Historical Checks vs Pipeline Baseline",
        "",
        "| Criterion | Historical pass rate | Tickets passing | Main comparison gap |",
        "|---|---:|---:|---|",
    ])
    for item in criterion_rows:
        lines.append(
            f"| {item['name']} | {item['pass_rate']}% | {item['passed']}/{item['total']} | {item['gap']} |"
        )

    lines.extend(_pipeline_comparison_lines(pipeline_comparisons, pipeline_limit))

    lines.extend([
        "",
        "## Ticket-Level Comparison",
        "",
        "| Ticket | Project | Score | Grade | Cases | Docs | Strengths | Gaps |",
        "|---|---|---:|---|---:|---|---|---|",
    ])
    for item in analyzed[:limit]:
        ticket = f"[{item.ticket_key}]({item.url})" if item.url else item.ticket_key
        lines.append(
            "| "
            + " | ".join(
                [
                    _md_cell(ticket),
                    _md_cell(item.project_key),
                    str(item.score),
                    item.grade,
                    str(item.test_case_count),
                    _md_cell(", ".join(item.requirement_docs) or "-"),
                    _md_cell("; ".join(item.strengths) or "-"),
                    _md_cell("; ".join(item.gaps) or "-"),
                ]
            )
            + " |"
        )

    lines.extend([
        "",
        "## Detailed Evidence",
        "",
    ])
    for item in analyzed[:limit]:
        lines.extend(_ticket_detail_lines(item))

    lines.extend([""])
    return "\n".join(lines)


def _pipeline_summary(comparisons: list[PipelineComparison]) -> dict[str, Any]:
    successful = [item for item in comparisons if not item.error]
    failed = [item for item in comparisons if item.error]
    generated_quality_rate = _pipeline_dimension_pass_rate(successful, _QUALITY_CRITERIA)
    generated_clarity_rate = _pipeline_dimension_pass_rate(successful, _CLARITY_CRITERIA)

    return {
        "attempted": len(comparisons),
        "successful": len(successful),
        "failed": len(failed),
        "avg_generated_cases": round(mean([item.generated_test_case_count for item in successful]), 1) if successful else 0,
        "avg_generated_score": round(mean([item.generated_score for item in successful]), 1) if successful else 0,
        "avg_score_delta": round(mean([item.score_delta for item in successful]), 1) if successful else 0,
        "avg_case_delta": round(mean([item.case_delta for item in successful]), 1) if successful else 0,
        "generated_quality_rate": generated_quality_rate,
        "generated_clarity_rate": generated_clarity_rate,
    }


def _pipeline_dimension_pass_rate(comparisons: list[PipelineComparison], criterion_names: set[str]) -> float:
    if not comparisons or not criterion_names:
        return 0
    total = 0
    passed = 0
    for item in comparisons:
        for name in item.generated_strengths:
            if name in criterion_names:
                passed += 1
        total += len(criterion_names)
    return round((passed / total) * 100, 1) if total else 0


def _pipeline_comparison_lines(
    comparisons: list[PipelineComparison],
    pipeline_limit: int,
) -> list[str]:
    lines = [
        "",
        "## Actual Pipeline Generation Comparison",
        "",
    ]
    if pipeline_limit <= 0:
        lines.append("Actual `/testcases/generate` comparison was disabled for this report.")
        return lines
    if not comparisons:
        lines.append("No filtered historical test-case tickets were available for actual pipeline generation.")
        return lines

    summary = _pipeline_summary(comparisons)
    lines.extend([
        f"Pipeline generation attempted for {summary['attempted']} filtered ticket(s), with {summary['successful']} success(es) and {summary['failed']} failure(s).",
        "Repository ownership is the `grounded_repos` value returned by `/testcases/generate`; when no repo override is supplied, RepoTree/Anthropic auto-detects the most relevant repository using repo summaries and Repomix directory previews before loading targeted Repomix context.",
        "",
        "| Ticket | Project | Detected repos | Existing cases | Generated cases | Existing score | Generated score | Score delta | Case delta | Repomix chars | Result |",
        "|---|---|---|---:|---:|---:|---:|---:|---:|---:|---|",
    ])
    for item in comparisons:
        result = item.error or "Generated"
        lines.append(
            "| "
            + " | ".join(
                [
                    _md_cell(item.ticket_key),
                    _md_cell(item.project_key),
                    _md_cell(", ".join(item.grounded_repos) or "-"),
                    str(item.historical_test_case_count),
                    str(item.generated_test_case_count),
                    str(item.historical_score),
                    str(item.generated_score),
                    str(item.score_delta),
                    str(item.case_delta),
                    str(item.repomix_context_chars),
                    _md_cell(result),
                ]
            )
            + " |"
        )

    lines.extend([
        "",
        "### Generated Output Evidence",
        "",
    ])
    for item in comparisons:
        lines.extend([
            f"#### {item.ticket_key}",
            "",
            f"- Detected repos: {', '.join(item.grounded_repos) or '-'}",
            f"- Semantic hits / files touched: {item.semantic_hits_count} / {item.files_touched_count}",
            f"- Existing vs generated score: {item.historical_score} -> {item.generated_score}",
        ])
        if item.error:
            lines.append(f"- Generation error: {_md_cell(item.error)}")
        else:
            lines.append(f"- Generated strengths: {', '.join(item.generated_strengths) or '-'}")
            lines.append(f"- Generated gaps: {', '.join(item.generated_gaps) or '-'}")
            if item.generated_snippet:
                lines.extend(["", item.generated_snippet])
        lines.append("")

    return lines


def _criterion_summary(analyzed: list[TicketQuality]) -> list[dict[str, Any]]:
    if not analyzed:
        return [
            {"name": name, "passed": 0, "total": 0, "pass_rate": 0, "gap": "No historical test cases found"}
            for name, _, _ in _rubric_rows()
        ]

    totals: dict[str, int] = Counter()
    passes: dict[str, int] = Counter()
    for ticket in analyzed:
        for criterion in ticket.criteria:
            totals[criterion.name] += 1
            if criterion.met:
                passes[criterion.name] += 1

    rows = []
    for name, _, expectation in _rubric_rows():
        total = totals.get(name, len(analyzed))
        passed = passes.get(name, 0)
        rate = round((passed / total) * 100, 1) if total else 0
        gap = "Aligned with pipeline" if rate >= 80 else expectation
        rows.append({"name": name, "passed": passed, "total": total, "pass_rate": rate, "gap": gap})
    return rows


def _ticket_detail_lines(item: TicketQuality) -> list[str]:
    lines = [
        f"### {item.ticket_key}: {item.summary}",
        "",
        f"- Project: {item.project_key}",
        f"- Status / Type: {item.status or '-'} / {item.issue_type or '-'}",
        f"- Score: {item.score}/100 ({item.grade})",
        f"- Estimated test cases: {item.test_case_count}",
        f"- Requirement docs found: {', '.join(item.requirement_docs) or '-'}",
        f"- Evidence sources: {', '.join(item.evidence_sources)}",
        "- Criteria:",
    ]
    for criterion in item.criteria:
        marker = "PASS" if criterion.met else "MISS"
        lines.append(f"  - {marker}: {criterion.name} ({criterion.weight}) - {criterion.detail}")
    if item.evidence_snippets:
        lines.append("- Evidence snippets:")
        for snippet in item.evidence_snippets:
            lines.append(f"  - {snippet}")
    lines.append("")
    return lines


def _rubric_rows() -> list[tuple[str, int, str]]:
    return [
        ("Identifiable test-case structure", 10, "Every case should have stable identifiers or scenario headings."),
        ("Step-by-step execution flow", 12, "Each case should have steps, numbered flow, or Gherkin Given/When/Then."),
        ("Expected results are stated", 12, "Each case should state one observable expected result."),
        ("Concrete assertions or verification points", 12, "Each case should include response, DB, side-effect, or UI assertions."),
        ("Coverage mix beyond happy path", 10, "Generated output should include happy, edge, and failure/error coverage."),
        ("Traceability to ticket or acceptance criteria", 10, "Cases should map back to ticket behavior or AC lines."),
        ("Test data, fixtures, and preconditions", 10, "Setup, payloads, mocks, DB fixtures, and auth fixtures should be explicit."),
        ("Source or implementation grounding", 12, "Generated cases should cite real source files, routes, services, tables, or snippets."),
        ("Automation readiness", 6, "Cases should say how they can be automated or executed."),
        ("Unambiguous expected contract", 6, "Expected results should avoid TBD, either/or, and vague outcomes."),
    ]


def _evidence_snippets(source_texts: list[tuple[str, str]], limit: int = 4) -> list[str]:
    snippets: list[str] = []
    for source, text in source_texts:
        clean = _normalize_space(text)
        if not clean:
            continue
        snippet = clean[:300]
        if len(clean) > 300:
            snippet += "..."
        snippets.append(f"{source}: {_md_cell(snippet)}")
        if len(snippets) >= limit:
            break
    return snippets


def _markdown_excerpt(value: str, limit: int = 1600) -> str:
    text = value.strip()
    if not text:
        return ""
    if len(text) > limit:
        text = text[:limit].rstrip() + "\n\n..."
    return text


def _has(pattern: str, text: str) -> bool:
    return re.search(pattern, text, flags=re.IGNORECASE | re.MULTILINE) is not None


def _grade(score: int) -> str:
    if score >= 80:
        return "High"
    if score >= 55:
        return "Medium"
    return "Low"


def _md_cell(value: str) -> str:
    return str(value).replace("|", "\\|").replace("\n", " ").strip()
