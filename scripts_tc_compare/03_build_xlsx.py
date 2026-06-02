"""Build the Test Case Comparative Analysis spreadsheet (.xlsx)."""
from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
IN_JSON = ROOT / "scripts_tc_compare" / "data_judged.json"
OUT_XLSX = ROOT / "scripts_tc_compare" / "Test_Case_Comparative_Analysis.xlsx"

HEADERS = [
    ("Ticket No", 14),
    ("Ticket Description", 55),
    ("Original Test Cases", 60),
    ("AI Gov Test Cases", 70),
    ("Original Score (/10)", 12),
    ("AI Gov Score (/10)", 12),
    ("Detailed Quality Difference", 75),
    ("CEO Summary", 55),
]


def main() -> None:
    records = json.loads(IN_JSON.read_text())

    wb = Workbook()
    ws = wb.active
    ws.title = "Comparative Analysis"

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    ws.append([h for h, _ in HEADERS])
    for idx, (_, width) in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(idx)].width = width

    for rec in records:
        ticket = rec.get("ticket_key", "")
        url = rec.get("url", "")
        ai_err = rec.get("ai_gov_error", "")
        ai_text = rec.get("ai_gov_test_cases") or (f"[generation error] {ai_err}" if ai_err else "")
        row = [
            ticket,
            rec.get("description", "") or rec.get("summary", ""),
            rec.get("original_test_cases", ""),
            ai_text,
            rec.get("original_score"),
            rec.get("ai_gov_score"),
            rec.get("detailed_difference", "") or rec.get("judge_error", ""),
            rec.get("ceo_summary", ""),
        ]
        ws.append(row)
        r = ws.max_row
        if url:
            cell = ws.cell(row=r, column=1)
            cell.hyperlink = url
            cell.font = Font(color="2563EB", underline="single")
        for col in range(1, len(HEADERS) + 1):
            ws.cell(row=r, column=col).alignment = Alignment(
                vertical="top", wrap_text=True, horizontal="center" if col in (1, 5, 6) else "left"
            )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Summary sheet
    s = wb.create_sheet("Summary", 0)
    scored = [r for r in records if isinstance(r.get("original_score"), (int, float))]
    avg_o = round(sum(r["original_score"] for r in scored) / len(scored), 2) if scored else 0
    avg_a = round(sum(r["ai_gov_score"] for r in scored) / len(scored), 2) if scored else 0
    gen_ok = sum(1 for r in records if r.get("ai_gov_test_cases") and not r.get("ai_gov_error"))
    rows = [
        ("Generated at", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
        ("Tickets compared", len(records)),
        ("AI-gov generations succeeded", f"{gen_ok}/{len(records)}"),
        ("Tickets judged", len(scored)),
        ("Average original score (/10)", avg_o),
        ("Average AI-gov score (/10)", avg_a),
        ("Average uplift (/10)", round(avg_a - avg_o, 2)),
        ("Judge model", "claude-opus-4-5 (fresh instance per ticket)"),
        ("AI governor", "RepoTree code-grounded test-case pipeline"),
    ]
    s.append(["Metric", "Value"])
    for cell in s[1]:
        cell.fill = header_fill
        cell.font = header_font
    for k, v in rows:
        s.append([k, v])
    s.column_dimensions["A"].width = 34
    s.column_dimensions["B"].width = 48

    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX}  (avg original {avg_o}/10, avg AI-gov {avg_a}/10)")


if __name__ == "__main__":
    main()
